"""Section 42: rehearse the organiser's validation experience end to end.

The organiser will not run our Makefile. They will open the Validation Lab,
drop in a spreadsheet we have never seen, and expect a scored file back. Every
way that can go wrong is a way we lose, and none of those ways show up in a
unit test - they show up in the shape of somebody else's export: columns in a
different order, a stray index column, resolution fields nobody stripped, a
category we never saw during training.

So this rehearses it against a running backend, over HTTP, the way they will.
The mock file is built from the locked test rows with the target removed, which
is the closest thing we have to data the model has never learned from.

Two invariants are checked after *every* variant, because they are the two that
would quietly turn a good score into a disqualification:

  * the bundle fingerprint is unchanged - scoring an upload must never write
    back into the model,
  * row order is preserved - a correct prediction against the wrong row is a
    wrong prediction.

Nothing here tunes anything. The labels are held out of the scoring frame by
the API itself and only opened afterwards, through the seal, so the accuracy
this prints was fixed before the first label was read (UPDATE 11 and 12).
"""
from __future__ import annotations

import datetime as dt
import hashlib
import io
import json
import sys
from typing import Any

import httpx
import numpy as np
import polars as pl

from muleguard import settings
from muleguard.data import ingest
from muleguard.data.split import load_locked_test_mask
from muleguard.logging import get_logger
from muleguard.utils import save_json

log = get_logger("cli.dry_run")

BASE = "http://127.0.0.1:8001"
MOCK_DIR = settings.ARTIFACTS_DIR / "dry_run"
MOCK_XLSX = MOCK_DIR / "mock_hidden_validation.xlsx"
LABELS_CSV = MOCK_DIR / "mock_hidden_labels.csv"
OUT = settings.METRICS_DIR / "organiser_dry_run.json"

TIMEOUT = 900.0

# Columns the firewall quarantines as post-decision. Two of the variants below
# deliberately put them back, because an organiser export very plausibly will.
RESOLUTION_COLS = ["F3898", "F3899", "F3913", "F3914", "F3915"]


# --- building the mock file ---------------------------------------------------

def _write_xlsx(df: pl.DataFrame, path) -> None:
    """Write a workbook with openpyxl, which is already a declared dependency.

    ``polars.write_excel`` wants xlsxwriter, and pulling in a package so a
    rehearsal can produce a file is a poor trade. Write-only mode streams rows
    instead of building the whole sheet in memory, which matters at ~7 million
    cells.
    """
    from openpyxl import Workbook

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("validation")
    ws.append(list(df.columns))
    for row in df.iter_rows():
        ws.append(list(row))
    wb.save(path)


def build_mock_file(rebuild: bool = False) -> dict[str, Any]:
    """Write `mock_hidden_validation.xlsx` from the locked test, target removed.

    The labels go to a *separate* CSV. Keeping them in a second file is not
    tidiness - it is the point. It makes "the scorer could not have seen the
    answer" a property of the filesystem rather than a claim in a report.
    """
    MOCK_DIR.mkdir(parents=True, exist_ok=True)
    df = ingest.load_dataset()
    mask = load_locked_test_mask()
    held = df.filter(pl.Series(mask))

    target = settings.TARGET_COLUMN
    labels = held[target].to_numpy()
    scoring = held.drop(target)

    if rebuild or not MOCK_XLSX.exists():
        _write_xlsx(scoring, MOCK_XLSX)
    scoring.write_csv(MOCK_DIR / "mock_hidden_validation.csv")
    pl.DataFrame({
        "row_order": np.arange(1, held.height + 1),
        target: labels,
    }).write_csv(LABELS_CSV)

    return {
        "path": str(MOCK_XLSX.relative_to(settings.REPO_ROOT)),
        "labels_path": str(LABELS_CSV.relative_to(settings.REPO_ROOT)),
        "rows": held.height,
        "columns": scoring.width,
        "target_removed": target,
        "positives_held_back": int(labels.sum()),
        "prevalence": round(float(labels.mean()), 5),
        "sha256": hashlib.sha256(MOCK_XLSX.read_bytes()).hexdigest()[:16],
        "provenance": "locked test split; no candidate model was fitted on these rows",
    }


# --- the six perturbations ----------------------------------------------------

def _required_features() -> set[str]:
    """The columns the champion actually needs, read from the bundle on disk.

    Read locally rather than over HTTP on purpose: `/v1/model` publishes counts
    but not the list, and widening a public route so a rehearsal can build a
    test file would be the tail wagging the dog.
    """
    from muleguard.models.scoring import load_bundle

    return set(load_bundle()["feature_list_selected"])


def _variants(df: pl.DataFrame) -> dict[str, pl.DataFrame]:
    """The malformations section 42 names, each built from the same base frame.

    Every one of these is something a real export does by accident. The names
    match the spec's list so the report can be read against it line by line.
    """
    rng = np.random.default_rng(settings.GLOBAL_SEED)

    shuffled = df.select(list(rng.permutation(df.columns)))

    extra = df.with_columns([
        pl.Series("analyst_notes", ["reviewed by ops"] * df.height),
        pl.Series("export_batch", ["BATCH-2026-08"] * df.height),
        pl.Series("Unnamed: 0", np.arange(df.height)),
    ])

    # F3912 is quarantined as post-decision, so its presence in an upload must
    # be ignored rather than consumed. Re-adding it with plausible values is the
    # only way to prove the firewall holds at inference and not just at fit.
    f3912 = df.with_columns(
        pl.Series("F3912", rng.integers(0, 3, df.height).astype(float)))

    resolution = df.with_columns([
        pl.Series(c, rng.normal(0, 1, df.height)) for c in RESOLUTION_COLS])

    # A category the encoder has never seen. The correct behaviour is to treat
    # it as unknown, not to crash and not to silently map it onto a known code.
    #
    # It has to land on a column the champion actually consumes, and on a
    # low-cardinality one, or the test proves nothing: an unseen value in a
    # column nobody reads is indistinguishable from no change at all. The first
    # version of this picked whichever column happened to parse as text, none of
    # which were selected features, and every variant came back byte-identical -
    # a green result that meant the probe had missed.
    required = _required_features()
    cat_col = next((c for c in df.columns
                    if c in required and df[c].n_unique() <= 20), None)
    if cat_col is None:
        cat_col = next(c for c in df.columns if c in required)
    categories = df.with_columns(pl.lit("NEVER_SEEN_CATEGORY").alias(cat_col))

    # Optional columns dropped. "Optional" is defined against the champion's
    # own feature list, not by position: dropping whatever happens to sit at
    # the end of the file would sometimes remove a required feature, and a
    # refusal would then be correct behaviour being scored as a failure.
    optional = [c for c in df.columns if c not in required][:40]
    missing = df.drop(optional)

    return {
        # The workbook itself, so the .xlsx reader is exercised on the file an
        # organiser actually sends; then the same frame as CSV, which is the
        # reference every other variant is compared against.
        "baseline": df,
        "baseline_csv": df,
        "shuffled_column_order": shuffled,
        "extra_columns": extra,
        "f3912_present": f3912,
        "resolution_fields_present": resolution,
        "category_changes": categories,
        "missing_optional_fields": missing,
        # Headers spelled as workbook variable names rather than F-numbers.
        # This must be *scored*, not refused: the data is correct and only the
        # spelling differs, so a refusal here would be a self-inflicted failure.
        "variable_name_headers": _variable_name_headers(df),
    }


def _variable_name_headers(df: pl.DataFrame) -> pl.DataFrame:
    """Rename what can be renamed to the workbook's human-readable names.

    Only names that resolve back unambiguously are used, so the variant tests
    header resolution rather than testing whether the file survives being made
    genuinely ambiguous.
    """
    from muleguard.validation.column_mapping import build_alias_index, resolve_columns

    _, varnames = build_alias_index()
    rename: dict[str, str] = {}
    for col in df.columns:
        vn = varnames.get(col)
        if not vn or vn.lower() == "nan" or vn in df.columns:
            continue
        # Round-trip: keep it only if the variable name maps back to this exact
        # feature. Anything ambiguous is left as its F-number.
        back = resolve_columns([vn])["mapping"].get(vn)
        if back == col and vn not in rename.values():
            rename[col] = vn
    return df.rename(rename) if rename else df


def _raw_variants(df: pl.DataFrame) -> dict[str, dict[str, Any]]:
    """Malformations that cannot be expressed as a valid DataFrame.

    These are the two files that must be **refused**. A corrupt workbook and a
    file with two columns of the same name are not cosmetic differences - one
    cannot be parsed at all and the other has no single correct interpretation.
    Scoring either would mean producing confident numbers from data we did not
    actually read, so a clean 4xx is the passing outcome here, not a failure.
    """
    csv_head = df.head(50)

    # A workbook truncated mid-stream: correct magic bytes, unreadable body.
    good = MOCK_XLSX.read_bytes()
    corrupted = good[: len(good) // 3]

    # Two columns with the same name. Written as raw CSV text because polars
    # will not construct such a frame - which is the point.
    cols = list(csv_head.columns[:20])
    dup_header = ",".join(cols + [cols[0]])
    lines = [dup_header]
    for row in csv_head.select(cols).iter_rows():
        vals = [("" if v is None else str(v)) for v in row]
        lines.append(",".join(vals + [vals[0]]))
    duplicate_csv = "\n".join(lines).encode("utf-8")

    return {
        "corrupted_xlsx": {
            "payload": corrupted,
            "filename": "corrupted.xlsx",
            "content_type": ("application/vnd.openxmlformats-officedocument"
                             ".spreadsheetml.sheet"),
            "must_be_refused": True,
            "why": ("a truncated workbook cannot be parsed; scoring it would "
                    "mean inventing rows"),
        },
        "duplicate_columns": {
            "payload": duplicate_csv,
            "filename": "duplicate_columns.csv",
            "content_type": "text/csv",
            "must_be_refused": True,
            "why": ("a repeated column name has no single correct reading, so "
                    "silently keeping one of the two would be a guess"),
        },
    }


# --- exercising the API -------------------------------------------------------

def _bundle_fingerprint() -> str:
    """A hash over the served bundle, to prove an upload did not modify it."""
    r = httpx.get(f"{BASE}/v1/model", timeout=60)
    r.raise_for_status()
    info = r.json()
    # The joblib sha is the strong part; the rest is here so that a bundle
    # reloaded with different thresholds or a different winner also shows up.
    keyed = {k: info.get(k) for k in
             ("bundle_sha256", "winner", "n_features_selected", "calibrator",
              "policy_thresholds", "data_fingerprint_sha256")}
    return hashlib.sha256(json.dumps(keyed, sort_keys=True, default=str)
                          .encode()).hexdigest()[:16]


def _upload(name: str, df: pl.DataFrame, as_xlsx: bool = False) -> dict[str, Any]:
    """POST one frame to the Lab.

    The baseline goes up as the real workbook so the xlsx path is exercised
    against the file an organiser would actually receive. The malformed
    variants go up as CSV: they differ from the baseline in *column shape*, not
    in container format, and both formats land in the same parser one line
    later - so re-encoding seven 7-million-cell workbooks would buy nothing but
    minutes.
    """
    if as_xlsx:
        payload, fname, ctype = (
            MOCK_XLSX.read_bytes(), f"{name}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        buf = io.BytesIO()
        df.write_csv(buf)
        payload, fname, ctype = buf.getvalue(), f"{name}.csv", "text/csv"
    r = httpx.post(f"{BASE}/v1/validation/run", timeout=TIMEOUT,
                   files={"file": (fname, payload, ctype)})
    return {"http_status": r.status_code,
            "body": r.json() if r.headers.get("content-type", "").startswith(
                "application/json") else {"text": r.text[:400]}}


def _check_variant(name: str, df: pl.DataFrame, fp_before: str) -> dict[str, Any]:
    """Run one variant and judge it on the four things that actually matter."""
    res = _upload(name, df, as_xlsx=(name == "baseline"))
    body, status = res["body"], res["http_status"]
    steps = {s["step"]: s for s in body.get("steps", [])} if status == 200 else {}
    seal = steps.get(3, {}).get("seal", {})

    checks = {
        # Every variant here is a file the Lab must *score*. Each one is a
        # legitimate export with a cosmetic difference, so there is no variant
        # in this list that a refusal would be the right answer to - which is
        # why none of them get an exemption.
        "accepted": status == 200,
        "schema_not_rejected": body.get("stopped_at_step") is None,
        # The mock workbook has no target, so the API must confirm it found
        # none and still scored. This is the "no target required" claim.
        "no_target_required": body.get("target_column_detected") is None,
        "scored": bool(body.get("seal_id")),
        "row_count_preserved": seal.get("n_rows") == df.height,
        "model_unchanged": _bundle_fingerprint() == fp_before,
    }
    return {
        "variant": name,
        "http_status": status,
        "overall": body.get("overall"),
        "seal_id": body.get("seal_id"),
        # The hash of the sealed prediction file. Two variants that share it
        # produced byte-identical predictions, which is how the invariance
        # claims below are settled without trusting a summary statistic.
        "prediction_sha256": seal.get("prediction_sha256"),
        "compatibility_score": body.get("compatibility_score"),
        "compatibility_band": body.get("compatibility_band"),
        "rows_in": df.height,
        "cols_in": df.width,
        "checks": checks,
        "verdict": "PASS" if all(checks.values()) else "FAIL",
        "failed_checks": [k for k, v in checks.items() if not v],
        # Carried even on success: a rejection with no recorded reason is a
        # failure nobody can diagnose six weeks later.
        "summary": body.get("summary") or body.get("next_action")
        or body.get("detail") or body.get("text"),
        "schema": {k: steps.get(1, {}).get(k) for k in
                   ("n_columns", "n_present", "n_missing_required",
                    "n_unexpected_columns", "schema_completeness")},
    }


def _check_raw_variant(name: str, spec: dict[str, Any],
                       fp_before: str) -> dict[str, Any]:
    """Run one malformed payload, where being refused is the passing outcome.

    Judged on three things: the Lab said no, it said no *cleanly* (a 4xx with a
    reason, not a 500 or a stack trace), and it did not score anything anyway.
    A crash is not a refusal - it is the same failure wearing a different code.
    """
    r = httpx.post(f"{BASE}/v1/validation/run", timeout=TIMEOUT,
                   files={"file": (spec["filename"], spec["payload"],
                                   spec["content_type"])})
    ctype = r.headers.get("content-type", "")
    body = r.json() if ctype.startswith("application/json") else {"text": r.text[:400]}
    status = r.status_code

    checks = {
        "refused": status >= 400,
        "refused_cleanly": 400 <= status < 500,
        "did_not_score": not body.get("seal_id"),
        "gave_a_reason": bool(body.get("detail") or body.get("summary")
                              or body.get("text")),
        "model_unchanged": _bundle_fingerprint() == fp_before,
    }
    return {
        "variant": name,
        "http_status": status,
        "expectation": "REFUSAL",
        "why_refusal_is_correct": spec["why"],
        "seal_id": body.get("seal_id"),
        "prediction_sha256": None,
        "compatibility_score": body.get("compatibility_score"),
        "compatibility_band": body.get("compatibility_band"),
        "checks": checks,
        "verdict": "PASS" if all(checks.values()) else "FAIL",
        "failed_checks": [k for k, v in checks.items() if not v],
        "summary": body.get("detail") or body.get("summary") or body.get("text"),
    }


def _reveal(seal_id: str) -> dict[str, Any]:
    """Step 8: compare against the held-back labels, after sealing.

    The labels travel as their own upload at reveal time. The endpoint verifies
    the prediction hash first and refuses if it moved, so a metric that comes
    back is a metric about predictions that predate the labels.
    """
    csv = LABELS_CSV.read_bytes()
    r = httpx.post(f"{BASE}/v1/validation/{seal_id}/reveal", timeout=TIMEOUT,
                   params={"label_column": settings.TARGET_COLUMN},
                   files={"file": ("labels.csv", csv, "text/csv")})
    if r.status_code != 200:
        return {"http_status": r.status_code, "detail": r.text[:400]}
    b = r.json()
    seal = b.get("seal", {})
    return {
        "http_status": 200,
        "state": seal.get("state"),
        "sealed_utc": seal.get("sealed_utc"),
        "revealed_utc": seal.get("revealed_utc"),
        "seal_verified": b.get("verification", {}).get("verified"),
        "prediction_sha256": seal.get("prediction_sha256"),
        "metrics": b.get("metrics"),
        "integrity_statement": b.get("integrity_statement"),
    }


# The variants that must not move a single prediction, and why. Two of these
# are the Feature Availability Firewall being tested where it counts: at
# inference, on a file that carries the quarantined columns anyway.
INVARIANT = {
    "shuffled_column_order": "column order is not information",
    "extra_columns": "columns the model does not consume are ignored",
    "f3912_present": "F3912 is quarantined, so supplying it must change nothing",
    "resolution_fields_present":
        "post-resolution fields are quarantined, so supplying them changes nothing",
    "missing_optional_fields": "columns the model does not consume are not needed",
}


def _invariance(results: list[dict[str, Any]]) -> dict[str, Any]:
    """Do the variants that should be indistinguishable have the same hash?

    The reference is `baseline_csv`, not `baseline`: the workbook and the CSV
    go through different readers, and a float that round-trips through xlsx can
    land one ulp away. Comparing like with like keeps this test about the
    firewall instead of about float formatting.
    """
    by_name = {r["variant"]: r for r in results}
    ref = by_name.get("baseline_csv", {}).get("prediction_sha256")
    out = {}
    for name, why in INVARIANT.items():
        got = by_name.get(name, {}).get("prediction_sha256")
        out[name] = {"identical_to_baseline": bool(ref) and got == ref,
                     "reason_it_must_be": why}
    # The control on the whole table. If corrupting a column the model *does*
    # read leaves the predictions unchanged, then the five "same" rows above
    # mean nothing - they would only show that the probe never reached the
    # model. This row is what makes the others evidence.
    changed = by_name.get("category_changes", {}).get("prediction_sha256")
    sensitivity = {
        "variant": "category_changes",
        "differs_from_baseline": bool(ref) and changed != ref,
        "reason_it_must_differ": (
            "an unseen category in a selected feature is a real change to the "
            "model's input; identical predictions here would mean the harness "
            "is not exercising the scorer"),
    }
    return {
        "reference_variant": "baseline_csv",
        "reference_prediction_sha256": ref,
        "results": out,
        "all_invariant": all(v["identical_to_baseline"] for v in out.values()),
        "sensitivity_control": sensitivity,
        "sound": (all(v["identical_to_baseline"] for v in out.values())
                  and sensitivity["differs_from_baseline"]),
    }


def run() -> dict[str, Any]:
    try:
        httpx.get(f"{BASE}/health/ready", timeout=10).raise_for_status()
    except Exception as e:  # noqa: BLE001 - the message matters more than the type
        raise SystemExit(
            f"no backend on {BASE} ({e}). Start one with:\n"
            f"  .venv/Scripts/python.exe -m uvicorn muleguard.api.main:app "
            f"--host 127.0.0.1 --port 8001")

    mock = build_mock_file()
    log.info("mock hidden validation file: %d rows x %d cols, %d positives held back",
             mock["rows"], mock["columns"], mock["positives_held_back"])

    base_df = pl.read_excel(MOCK_XLSX)
    fp_before = _bundle_fingerprint()

    results = []
    for name, df in _variants(base_df).items():
        r = _check_variant(name, df, fp_before)
        r["expectation"] = "SCORED"
        results.append(r)
        log.info("variant %-28s %s (compat %s)", name, r["verdict"],
                 r.get("compatibility_score"))

    # Malformed payloads, judged on being refused rather than on being scored.
    for name, spec in _raw_variants(base_df).items():
        r = _check_raw_variant(name, spec, fp_before)
        results.append(r)
        log.info("variant %-28s %s (refusal expected, got %s)", name,
                 r["verdict"], r["http_status"])

    baseline = next(r for r in results if r["variant"] == "baseline")
    reveal = _reveal(baseline["seal_id"]) if baseline.get("seal_id") else {
        "skipped": "baseline produced no seal"}

    fp_after = _bundle_fingerprint()
    passed = [r for r in results if r["verdict"] == "PASS"]
    inv = _invariance(results)

    payload = {
        "generated_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
        "spec": "section 42 - dry-run the organiser validation experience",
        "mock_file": mock,
        "bundle_fingerprint_before": fp_before,
        "bundle_fingerprint_after": fp_after,
        "accepted_model_unchanged": fp_before == fp_after,
        "variants": results,
        "variants_passed": f"{len(passed)}/{len(results)}",
        "prediction_invariance": inv,
        "offline_label_comparison": reveal,
        "leakage_statement": (
            "the revealed metric cannot have leaked into training: the target "
            "column was dropped from the uploaded workbook before it left this "
            "process, the API withholds any label column it finds before "
            "scoring, and the labels were supplied only at reveal time against "
            "a prediction file whose hash was already recorded"),
        "verdict": "PASS" if (fp_before == fp_after
                              and len(passed) == len(results)
                              and inv["sound"]) else "FAIL",
    }
    save_json(payload, OUT)
    return payload


if __name__ == "__main__":
    p = run()
    print(f"\nmock file  {p['mock_file']['rows']} rows x {p['mock_file']['columns']} cols, "
          f"{p['mock_file']['positives_held_back']} positives withheld")
    print(f"{'variant':<30} {'status':>6}  {'compat':>6}  verdict")
    for r in p["variants"]:
        c = r["compatibility_score"]
        print(f"{r['variant']:<30} {r['http_status']:>6}  "
              f"{c if c is not None else '-':>6}  {r['verdict']}")
    print(f"\nprediction invariance vs {p['prediction_invariance']['reference_variant']}:")
    for name, v in p["prediction_invariance"]["results"].items():
        print(f"  {'same' if v['identical_to_baseline'] else 'DIFFERS':>7}  "
              f"{name:<28} {v['reason_it_must_be']}")
    sc = p["prediction_invariance"]["sensitivity_control"]
    print(f"  {'DIFFERS' if sc['differs_from_baseline'] else 'same (BAD)':>7}  "
          f"{sc['variant']:<28} control: proves the scorer was reached")
    print(f"\naccepted model unchanged: {p['accepted_model_unchanged']} "
          f"({p['bundle_fingerprint_before']} -> {p['bundle_fingerprint_after']})")
    m = p["offline_label_comparison"].get("metrics") or {}
    if m:
        print("offline comparison:", {k: m[k] for k in list(m)[:6]})
    print(f"\nDRY RUN {p['verdict']}  ({p['variants_passed']} variants)")
    sys.exit(0 if p["verdict"] == "PASS" else 1)
