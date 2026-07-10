"""Immutable raw-data handling and one-time XLSX -> Parquet conversion.

Rules enforced here (master prompt section 8):
- the original workbook is never modified; a read-only copy lives in data/raw/
- conversion happens once, to data/interim/dataset.parquet (+ schema, fingerprint)
- the conversion is validated against an INDEPENDENT engine (openpyxl streaming)
  on: exact row count, exact column count, column-name equality, target
  distribution, and random-cell spot checks.
"""
from __future__ import annotations

import datetime as dt
import math
import os
import stat
from pathlib import Path
from typing import Any

import fastexcel
import numpy as np
import polars as pl

from muleguard import settings
from muleguard.logging import get_logger
from muleguard.utils import save_json, sha256_file

log = get_logger("data.ingest")


def ensure_raw_copy() -> dict[str, Any]:
    """Copy the original workbook into data/raw/ (read-only) and hash both."""
    cfg = settings.load_config("data")
    original = settings.REPO_ROOT / cfg["raw"]["original_xlsx"]
    raw_copy = settings.REPO_ROOT / cfg["raw"]["raw_copy"]
    if not original.exists():
        raise FileNotFoundError(f"Original dataset not found: {original}")

    raw_copy.parent.mkdir(parents=True, exist_ok=True)
    if not raw_copy.exists():
        # copy bytes, then drop write permission on the copy
        raw_copy.write_bytes(original.read_bytes())
        os.chmod(raw_copy, stat.S_IREAD)
        log.info("raw copy created at %s (read-only)", raw_copy)

    sha_original = sha256_file(original)
    sha_copy = sha256_file(raw_copy)
    if sha_original != sha_copy:
        raise RuntimeError("raw copy hash mismatch - data/raw copy is corrupt")
    return {
        "original_path": str(original),
        "raw_copy_path": str(raw_copy),
        "sha256": sha_copy,
        "size_bytes": raw_copy.stat().st_size,
    }


def convert_to_parquet(force: bool = False) -> dict[str, Any]:
    """Read the raw XLSX once (calamine engine) and write Parquet + schema."""
    cfg = settings.load_config("data")
    raw_copy = settings.REPO_ROOT / cfg["raw"]["raw_copy"]
    parquet_path = settings.REPO_ROOT / cfg["interim"]["parquet"]
    schema_path = settings.REPO_ROOT / cfg["interim"]["schema_json"]

    if parquet_path.exists() and not force:
        log.info("parquet already exists, skipping conversion: %s", parquet_path)
        return {"parquet_path": str(parquet_path), "skipped": True}

    reader = fastexcel.read_excel(str(raw_copy))
    sheet_names = reader.sheet_names
    sheet = reader.load_sheet(cfg["raw"]["sheet_index"])
    df: pl.DataFrame = sheet.to_polars()
    log.info("loaded sheet %r: %d rows x %d cols", sheet_names[cfg["raw"]["sheet_index"]], df.height, df.width)

    if df.width != len(set(df.columns)):
        raise RuntimeError("duplicate column names detected in raw workbook")

    parquet_path.parent.mkdir(parents=True, exist_ok=True)
    df.write_parquet(parquet_path, compression="zstd")

    schema = {
        "sheet_names": sheet_names,
        "used_sheet_index": cfg["raw"]["sheet_index"],
        "n_rows": df.height,
        "n_cols": df.width,
        "columns": [{"name": c, "dtype": str(df.schema[c])} for c in df.columns],
    }
    save_json(schema, schema_path)
    return {
        "parquet_path": str(parquet_path),
        "schema_path": str(schema_path),
        "n_rows": df.height,
        "n_cols": df.width,
        "skipped": False,
    }


def load_raw_parquet() -> pl.DataFrame:
    cfg = settings.load_config("data")
    return pl.read_parquet(settings.REPO_ROOT / cfg["interim"]["parquet"])


# Date formats observed in F3888 (account opening date): true date cells are
# rendered ISO by calamine; text cells use month-day-year with hyphens.
_F3888_FORMATS = ["%Y-%m-%d %H:%M:%S", "%m-%d-%Y", "%Y-%m-%d", "%d-%m-%Y"]


def canonicalize(df: pl.DataFrame) -> tuple[pl.DataFrame, dict[str, Any]]:
    """Deterministic, target-blind canonicalisation of the raw parquet.

    - the literal string ``"NA"`` is the workbook's missing marker -> null
    - string columns whose non-NA values all parse as numbers -> Float64
    - ``F3888`` (opening date, mixed text formats) -> Date
    - remaining string columns stay Utf8 (true categoricals)
    - datetime columns pass through unchanged

    Returns the canonical frame and a parse report (per-column NA counts and
    parse failures; any unexpected failure is loud in the report).
    """
    report: dict[str, Any] = {"na_string_nulls": {}, "parse_failures": {}, "coerced_numeric": [],
                              "parsed_dates": [], "kept_categorical": []}
    exprs: list[pl.Expr] = []
    for c in df.columns:
        dt_ = df.schema[c]
        if dt_.is_numeric() or dt_.is_temporal():
            exprs.append(pl.col(c))
            continue
        s = df[c].cast(pl.Utf8).str.strip_chars()
        na_count = int((s == "NA").sum())
        cleaned = pl.col(c).cast(pl.Utf8).str.strip_chars()
        cleaned = pl.when(cleaned == "NA").then(None).otherwise(cleaned)
        non_na = s.filter((s != "NA") & s.is_not_null())
        as_num = non_na.cast(pl.Float64, strict=False)
        n_unparseable = int(as_num.null_count())
        if na_count:
            report["na_string_nulls"][c] = na_count
        if c == "F3888":
            parsed = None
            for fmt in _F3888_FORMATS:
                trial = non_na.str.strptime(pl.Date, fmt, strict=False)
                parsed = trial if parsed is None else parsed.fill_null(trial)
            date_expr = None
            for fmt in _F3888_FORMATS:
                e = cleaned.str.strptime(pl.Date, fmt, strict=False)
                date_expr = e if date_expr is None else date_expr.fill_null(e)
            fails = int(parsed.null_count()) if parsed is not None else len(non_na)
            report["parse_failures"][c] = fails
            report["parsed_dates"].append(c)
            exprs.append(date_expr.alias(c))
        elif n_unparseable == 0:
            report["coerced_numeric"].append(c)
            exprs.append(cleaned.cast(pl.Float64, strict=False).alias(c))
        else:
            report["kept_categorical"].append(c)
            exprs.append(cleaned.alias(c))
    out = df.select(exprs)
    return out, report


def load_dataset() -> pl.DataFrame:
    """Canonical analysis dataset (cached parquet next to the raw one)."""
    cfg = settings.load_config("data")
    canon_path = settings.REPO_ROOT / cfg["interim"]["parquet"]
    canon_path = canon_path.with_name("dataset_canonical.parquet")
    if canon_path.exists():
        return pl.read_parquet(canon_path)
    df, report = canonicalize(load_raw_parquet())
    df.write_parquet(canon_path, compression="zstd")
    save_json(report, canon_path.with_name("canonicalization_report.json"))
    return df


def _parses_close(text: str, number: float) -> bool:
    try:
        return math.isclose(float(text.strip()), number, rel_tol=1e-9, abs_tol=1e-12)
    except (ValueError, AttributeError):
        return False


def _spot_check_targets(n_rows: int, n_cols: int, n_cells: int, seed: int) -> list[tuple[int, int]]:
    rng = np.random.default_rng(seed)
    rows = rng.integers(0, n_rows, size=n_cells)
    cols = rng.integers(0, n_cols, size=n_cells)
    return sorted(zip(rows.tolist(), cols.tolist()))


def validate_against_independent_engine() -> dict[str, Any]:
    """Stream the workbook with openpyxl (independent engine) and compare.

    Single lazy pass computing: row/col counts, header equality, target
    distribution and randomly pre-selected spot cells.
    """
    import openpyxl  # local import: heavy

    cfg = settings.load_config("data")
    raw_copy = settings.REPO_ROOT / cfg["raw"]["raw_copy"]
    df = load_raw_parquet()  # validates the CONVERSION; canonicalisation has its own report
    n_cells = int(cfg["audit"]["spot_check_cells"])
    targets = _spot_check_targets(df.height, df.width, n_cells, settings.GLOBAL_SEED)
    wanted_rows: dict[int, list[int]] = {}
    for r, c in targets:
        wanted_rows.setdefault(r, []).append(c)

    wb = openpyxl.load_workbook(raw_copy, read_only=True, data_only=True)
    ws = wb.worksheets[cfg["raw"]["sheet_index"]]

    header: list[Any] | None = None
    n_rows = 0
    target_counts: dict[Any, int] = {}
    spot_values: dict[tuple[int, int], Any] = {}
    target_idx: int | None = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = list(row)
            try:
                target_idx = header.index(settings.TARGET_COLUMN)
            except ValueError:
                target_idx = None
            continue
        data_row_idx = i - 1
        n_rows += 1
        if target_idx is not None:
            tv = row[target_idx] if target_idx < len(row) else None
            target_counts[tv] = target_counts.get(tv, 0) + 1
        if data_row_idx in wanted_rows:
            for c in wanted_rows[data_row_idx]:
                spot_values[(data_row_idx, c)] = row[c] if c < len(row) else None
    wb.close()

    # --- compare ---
    problems: list[str] = []
    if header is None:
        problems.append("workbook appears empty")
        header = []
    raw_header = [h if h is not None else "" for h in header]
    pl_header = df.columns
    # fastexcel names unnamed columns; compare named columns strictly and
    # record any unnamed-header positions explicitly.
    unnamed_positions = [i for i, h in enumerate(raw_header) if str(h).strip() == ""]
    named_mismatches = [
        (i, raw, got)
        for i, (raw, got) in enumerate(zip(raw_header, pl_header))
        if str(raw).strip() != "" and str(raw) != got
    ]
    if named_mismatches:
        problems.append(f"column-name mismatches: {named_mismatches[:5]}")
    if n_rows != df.height:
        problems.append(f"row count mismatch: openpyxl={n_rows} parquet={df.height}")
    if len(raw_header) != df.width:
        problems.append(f"col count mismatch: openpyxl={len(raw_header)} parquet={df.width}")

    tgt_parquet = (
        df[settings.TARGET_COLUMN].value_counts().sort(settings.TARGET_COLUMN)
        if settings.TARGET_COLUMN in df.columns else None
    )
    tgt_parquet_dict = (
        {row[0]: row[1] for row in tgt_parquet.iter_rows()} if tgt_parquet is not None else {}
    )
    norm_openpyxl = {(int(k) if k is not None else None): v for k, v in target_counts.items()}
    norm_parquet = {(int(k) if k is not None else None): v for k, v in tgt_parquet_dict.items()}
    if norm_openpyxl != norm_parquet:
        problems.append(
            f"target distribution mismatch: openpyxl={norm_openpyxl} parquet={norm_parquet}"
        )

    mismatched_cells = []
    for (r, c), raw_val in spot_values.items():
        pq_val = df[r, c]
        # Workbook convention: the literal string "NA" is the missing marker.
        # calamine maps it to null in numeric-inferred columns; both readings
        # denote the same missing cell.
        raw_is_na = raw_val is None or (isinstance(raw_val, str) and raw_val.strip() == "NA")
        pq_is_na = pq_val is None or (isinstance(pq_val, str) and pq_val.strip() == "NA")
        ok = (
            (raw_is_na and pq_is_na)
            or (raw_val == pq_val)
            or (
                isinstance(raw_val, (int, float))
                and isinstance(pq_val, (int, float))
                and math.isclose(float(raw_val), float(pq_val), rel_tol=1e-9, abs_tol=1e-12)
            )
            # numeric stored as text in one engine, number in the other
            or (
                isinstance(raw_val, str) and isinstance(pq_val, (int, float))
                and _parses_close(raw_val, float(pq_val))
            )
            or (
                isinstance(pq_val, str) and isinstance(raw_val, (int, float))
                and _parses_close(pq_val, float(raw_val))
            )
            # datetime cell vs its string rendering
            or (str(raw_val) == str(pq_val))
            or (
                isinstance(raw_val, dt.datetime)
                and str(pq_val).startswith(str(raw_val.date()))
            )
        )
        if not ok:
            mismatched_cells.append({"row": r, "col": c, "raw": raw_val, "parquet": pq_val})
    if mismatched_cells:
        problems.append(f"spot-check mismatches: {mismatched_cells}")

    result = {
        "validated_at_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
        "engine": "openpyxl read-only streaming",
        "n_rows_raw": n_rows,
        "n_cols_raw": len(raw_header),
        "unnamed_header_positions": unnamed_positions,
        "target_distribution_raw": {str(k): v for k, v in norm_openpyxl.items()},
        "spot_cells_checked": len(spot_values),
        "problems": problems,
        "passed": not problems,
    }
    return result


def build_fingerprint(raw_info: dict[str, Any], validation: dict[str, Any]) -> dict[str, Any]:
    """Assemble and persist the immutable dataset fingerprint."""
    cfg = settings.load_config("data")
    parquet_path = settings.REPO_ROOT / cfg["interim"]["parquet"]
    df = load_raw_parquet()
    tgt = df[settings.TARGET_COLUMN]
    n_pos = int((tgt == 1).sum())
    n_neg = int((tgt == 0).sum())
    n_null = int(tgt.null_count())
    fingerprint = {
        "created_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
        "raw_file": raw_info,
        "parquet_file": {
            "path": str(parquet_path),
            "sha256": sha256_file(parquet_path),
            "size_bytes": parquet_path.stat().st_size,
        },
        "n_rows": df.height,
        "n_cols": df.width,
        "target_column": settings.TARGET_COLUMN,
        "target_distribution": {"positives_1": n_pos, "negatives_0": n_neg, "null": n_null},
        "positive_prevalence": round(n_pos / max(n_pos + n_neg, 1), 6),
        "columns_sha256": sha256_file_columns(df),
        "independent_validation": validation,
    }
    save_json(fingerprint, settings.REPO_ROOT / cfg["interim"]["fingerprint_json"])
    return fingerprint


def sha256_file_columns(df: pl.DataFrame) -> str:
    """Hash the ordered column-name list (schema identity)."""
    import hashlib

    h = hashlib.sha256()
    for c in df.columns:
        h.update(c.encode("utf-8"))
        h.update(b"\x00")
    return h.hexdigest()
